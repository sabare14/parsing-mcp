"""Deterministic Excel template detection with additive, inspectable scoring.
This module detects:
- best sheet
- header row
- first data row
Design goals:
- compact feature extraction
- centralized weights
- purely additive scoring
- full score breakdowns for LLM-guided tuning
"""
from __future__ import annotations
import copy
import os
import re
import zipfile
from io import BytesIO
from typing import Any
from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl.worksheet.worksheet import Worksheet
SUPPORTED_EXCEL_EXTENSIONS: tuple[str, ...] = (".xlsx", ".xlsm")
MAX_SCAN_ROWS_DEFAULT = 15
MAX_SCAN_COLUMNS = 200
TOP_K_DEFAULT = 3
SHEET_NAME_POSITIVE_TOKENS = ("template", "data", "input", "upload")
SHEET_NAME_NEGATIVE_TOKENS = (
    "instruction",
    "readme",
    "note",
    "example",
    "sample",
    "lookup",
    "lov",
    "list",
    "ref",
)
TOKEN_SIGNALS = {
    "id_like": {
        "id",
        "ids",
        "uid",
        "uuid",
        "itemid",
        "productid",
        "ref",
        "reference",
        "serial",
        "number",
        "no",
    },
    "name_like": {
        "name",
        "title",
        "product",
        "item",
        "brand",
        "description",
        "desc",
    },
    "date_like": {
        "date",
        "day",
        "month",
        "year",
        "dob",
        "expiry",
        "validity",
        "from",
        "to",
        "dd",
        "mm",
        "yyyy",
    },
    "code_like": {
        "code",
        "sku",
        "hsn",
        "ean",
        "upc",
        "barcode",
        "pin",
        "pincode",
        "zip",
        "style",
    },
}
WEIGHTS = {
    "sheet": {
        "column_breadth": 0.32,
        "density": 0.04,
        "text_balance": 0.10,
        "style_support": 0.05,
        "name_positive": 0.10,
        "template_shape": 0.16,
        "header_focus": 0.05,
        "wide_sparse_ratio": 0.08,
        "instruction_signal": -0.22,
        "lookup_signal": -0.24,
        "dense_penalty": -0.06,
        "name_negative": -0.45,
    },
    "header": {
        "non_empty_norm": 0.22,
        "string_ratio": 0.18,
        "short_text_ratio": 0.13,
        "unique_value_ratio": 0.08,
        "style_ratio": 0.08,
        "followed_by_tabular": 0.12,
        "early_row_bias": 0.02,
        "has_id_like_token": 0.05,
        "has_name_like_token": 0.07,
        "has_date_like_token": 0.05,
        "has_code_like_token": 0.05,
        "numeric_ratio": -0.16,
        "long_text_ratio": -0.10,
        "sparse_penalty": -0.12,
    },
    "data": {
        "overlap_with_header": 0.20,
        "numeric_ratio": 0.08,
        "value_mix": 0.10,
        "early_after_header": 0.12,
        "offset_preference": 0.12,
        "unstyled_ratio": 0.05,
        "unique_value_ratio": 0.06,
        "first_non_empty_after_header": 0.14,
        "sparse_input": 0.06,
        "transition_from_dense": 0.12,
        "header_like_text": -0.22,
        "long_note_penalty": -0.14,
        "dense_row_penalty": -0.06,
    },
    "selection": {
        "sheet": 0.80,
        "header": 0.15,
        "data": 0.05,
    },
    "workbook": {
        "best_sheet_selection": 0.90,
        "visible_sheet_bonus": 0.10,
    },
    "confidence": {
        "workbook": 0.20,
        "sheet": 0.30,
        "header": 0.30,
        "data": 0.20,
    },
}
def clamp(value: float, low: float = 0.0, high: float = 1.0) -> float:
    return max(low, min(high, float(value)))
def _tokenize(text: str) -> list[str]:
    return re.findall(r"[a-z0-9]+", text.lower())
def _resolve_settings(max_rows: int, knobs: dict[str, Any] | None) -> dict[str, Any]:
    settings = {
        "max_scan_rows": max(1, int(max_rows)),
        "max_scan_columns": MAX_SCAN_COLUMNS,
        "top_k": TOP_K_DEFAULT,
        "weights": copy.deepcopy(WEIGHTS),
        "sheet_name_positive_tokens": SHEET_NAME_POSITIVE_TOKENS,
        "sheet_name_negative_tokens": SHEET_NAME_NEGATIVE_TOKENS,
    }
    if not isinstance(knobs, dict):
        return settings
    if "max_scan_rows_default" in knobs:
        settings["max_scan_rows"] = max(1, int(knobs["max_scan_rows_default"]))
    if "max_scan_rows" in knobs:
        settings["max_scan_rows"] = max(1, int(knobs["max_scan_rows"]))
    if "max_scan_columns" in knobs:
        settings["max_scan_columns"] = max(1, int(knobs["max_scan_columns"]))
    if "top_k" in knobs:
        settings["top_k"] = max(1, int(knobs["top_k"]))
    if isinstance(knobs.get("sheet_name_positive_tokens"), (list, tuple)):
        settings["sheet_name_positive_tokens"] = tuple(
            str(x).strip().lower() for x in knobs["sheet_name_positive_tokens"] if str(x).strip()
        )
    if isinstance(knobs.get("sheet_name_negative_tokens"), (list, tuple)):
        settings["sheet_name_negative_tokens"] = tuple(
            str(x).strip().lower() for x in knobs["sheet_name_negative_tokens"] if str(x).strip()
        )
    if isinstance(knobs.get("weights"), dict):
        for section, section_weights in knobs["weights"].items():
            if section in settings["weights"] and isinstance(section_weights, dict):
                for key, value in section_weights.items():
                    if key in settings["weights"][section]:
                        settings["weights"][section][key] = float(value)
    legacy_conf_keys = {
        "workbook_conf_weight": "workbook",
        "sheet_conf_weight": "sheet",
        "header_conf_weight": "header",
        "data_conf_weight": "data",
    }
    if any(key in knobs for key in legacy_conf_keys):
        confidence = settings["weights"]["confidence"]
        for knob_key, conf_key in legacy_conf_keys.items():
            if knob_key in knobs:
                confidence[conf_key] = float(knobs[knob_key])
        total = sum(max(0.0, v) for v in confidence.values())
        if total > 0:
            for key, value in confidence.items():
                confidence[key] = max(0.0, value) / total
    return settings
def extract_excel_files(zip_path: str) -> list[tuple[str, bytes]]:
    """Read supported Excel files from a ZIP archive."""
    try:
        with zipfile.ZipFile(zip_path, "r") as archive:
            names = sorted(
                name
                for name in archive.namelist()
                if not name.endswith("/") and name.lower().endswith(SUPPORTED_EXCEL_EXTENSIONS)
            )
            if not names:
                supported = ", ".join(SUPPORTED_EXCEL_EXTENSIONS)
                raise ValueError(f"No supported Excel files ({supported}) found in ZIP archive.")
            return [(name, archive.read(name)) for name in names]
    except zipfile.BadZipFile as exc:
        raise ValueError(f"Invalid ZIP file: {zip_path}") from exc
    except FileNotFoundError as exc:
        raise ValueError(f"Invalid ZIP file: {zip_path}") from exc
def load_workbook(xlsx_bytes: bytes) -> Any:
    """Load an openpyxl workbook from bytes."""
    try:
        return openpyxl_load_workbook(filename=BytesIO(xlsx_bytes), data_only=False, read_only=False)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Failed to load workbook: {exc}") from exc
def get_visible_sheets(workbook: Any) -> list[Worksheet]:
    return [ws for ws in workbook.worksheets if getattr(ws, "sheet_state", "visible") == "visible"]
def _extract_row_features(ws: Worksheet, row_idx: int, max_columns: int) -> dict[str, Any]:
    row_cells = list(
        ws.iter_rows(min_row=row_idx, max_row=row_idx, min_col=1, max_col=max_columns, values_only=False)
    )
    if not row_cells:
        return {
            "row_index": row_idx,
            "non_empty_count": 0,
            "string_count": 0,
            "numeric_count": 0,
            "string_ratio": 0.0,
            "numeric_ratio": 0.0,
            "short_text_ratio": 0.0,
            "long_text_ratio": 0.0,
            "average_text_length": 0.0,
            "unique_value_ratio": 0.0,
            "styled_count": 0,
            "bold_count": 0,
            "non_empty_columns": [],
            "has_id_like_token": False,
            "has_name_like_token": False,
            "has_date_like_token": False,
            "has_code_like_token": False,
            "token_samples": {"id_like": [], "name_like": [], "date_like": [], "code_like": []},
        }
    non_empty_count = 0
    string_count = 0
    numeric_count = 0
    short_text_count = 0
    long_text_count = 0
    styled_count = 0
    bold_count = 0
    text_lengths: list[int] = []
    unique_values: set[str] = set()
    non_empty_columns: list[int] = []
    token_hits = {"id_like": False, "name_like": False, "date_like": False, "code_like": False}
    token_samples = {"id_like": [], "name_like": [], "date_like": [], "code_like": []}
    for col_idx, cell in enumerate(row_cells[0], start=1):
        value = cell.value
        if value is None or (isinstance(value, str) and not value.strip()):
            continue
        non_empty_count += 1
        non_empty_columns.append(col_idx)
        is_bold = bool(getattr(getattr(cell, "font", None), "bold", False))
        fill = getattr(cell, "fill", None)
        has_fill = bool(fill and getattr(fill, "fill_type", None) not in (None, "none"))
        if is_bold:
            bold_count += 1
        if is_bold or has_fill:
            styled_count += 1
        if isinstance(value, bool):
            unique_values.add(str(value).lower())
            continue
        if isinstance(value, (int, float)):
            numeric_count += 1
            unique_values.add(str(value))
            continue
        text = str(value).strip()
        if not text:
            continue
        string_count += 1
        unique_values.add(text.lower())
        text_lengths.append(len(text))
        if len(text) <= 30:
            short_text_count += 1
        if len(text) >= 45:
            long_text_count += 1
        for token in _tokenize(text):
            for signal_name, keywords in TOKEN_SIGNALS.items():
                if token in keywords:
                    token_hits[signal_name] = True
                    samples = token_samples[signal_name]
                    if token not in samples and len(samples) < 3:
                        samples.append(token)
    string_ratio = string_count / non_empty_count if non_empty_count else 0.0
    numeric_ratio = numeric_count / non_empty_count if non_empty_count else 0.0
    short_text_ratio = short_text_count / string_count if string_count else 0.0
    long_text_ratio = long_text_count / string_count if string_count else 0.0
    average_text_length = (sum(text_lengths) / len(text_lengths)) if text_lengths else 0.0
    unique_value_ratio = len(unique_values) / non_empty_count if non_empty_count else 0.0
    return {
        "row_index": row_idx,
        "non_empty_count": non_empty_count,
        "string_count": string_count,
        "numeric_count": numeric_count,
        "string_ratio": string_ratio,
        "numeric_ratio": numeric_ratio,
        "short_text_ratio": short_text_ratio,
        "long_text_ratio": long_text_ratio,
        "average_text_length": average_text_length,
        "unique_value_ratio": unique_value_ratio,
        "styled_count": styled_count,
        "bold_count": bold_count,
        "non_empty_columns": non_empty_columns,
        "has_id_like_token": token_hits["id_like"],
        "has_name_like_token": token_hits["name_like"],
        "has_date_like_token": token_hits["date_like"],
        "has_code_like_token": token_hits["code_like"],
        "token_samples": token_samples,
    }
def extract_features(
    ws: Worksheet,
    max_rows: int,
    max_columns: int,
    positive_name_tokens: tuple[str, ...],
    negative_name_tokens: tuple[str, ...],
) -> dict[str, Any]:
    """Extract sheet-level and row-level features for additive scoring."""
    scanned_rows = max(1, int(max_rows))
    scanned_columns = min(max_columns, max(1, int(getattr(ws, "max_column", 1) or 1)))
    rows = [_extract_row_features(ws, row_idx, scanned_columns) for row_idx in range(1, scanned_rows + 1)]
    total_non_empty = sum(row["non_empty_count"] for row in rows)
    total_strings = sum(row["string_count"] for row in rows)
    total_numeric = sum(row["numeric_count"] for row in rows)
    total_styled = sum(row["styled_count"] for row in rows)
    active_columns = sorted({col for row in rows for col in row["non_empty_columns"]})
    active_column_count = len(active_columns)
    average_non_empty_per_row = total_non_empty / scanned_rows if scanned_rows else 0.0
    text_cell_ratio = total_strings / total_non_empty if total_non_empty else 0.0
    numeric_cell_ratio = total_numeric / total_non_empty if total_non_empty else 0.0
    styled_cell_ratio = total_styled / total_non_empty if total_non_empty else 0.0
    header_like_row_ratio = (
        sum(
            1
            for row in rows
            if row["non_empty_count"] >= 3
            and row["string_ratio"] >= 0.6
            and row["short_text_ratio"] >= 0.5
        )
        / scanned_rows
    )
    sparse_row_ratio = sum(1 for row in rows if row["non_empty_count"] <= 1) / scanned_rows
    instruction_signal = (
        sum(1 for row in rows if row["non_empty_count"] <= 2 and row["long_text_ratio"] >= 0.5) / scanned_rows
    )
    lookup_signal = clamp(
        clamp((10.0 - active_column_count) / 10.0)
        * clamp(average_non_empty_per_row / 4.0)
        * clamp((numeric_cell_ratio + 0.4 * text_cell_ratio) / 0.8)
    )
    candidate_header_rows = [
        row["row_index"]
        for row in rows
        if row["non_empty_count"] >= 3 and row["string_ratio"] >= 0.6 and row["short_text_ratio"] >= 0.5
    ]
    first_header_row = candidate_header_rows[0] if candidate_header_rows else (scanned_rows + 1)
    trailing_rows = [row for row in rows if row["row_index"] > first_header_row]
    trailing_sparse_ratio = (
        sum(1 for row in trailing_rows if row["non_empty_count"] <= 1) / len(trailing_rows)
        if trailing_rows
        else 0.0
    )
    early_header_signal = clamp((4.0 - first_header_row) / 3.0) if first_header_row <= 4 else 0.0
    template_shape = clamp(0.65 * trailing_sparse_ratio + 0.35 * early_header_signal)
    header_focus = clamp(1.0 - 1.4 * header_like_row_ratio)
    wide_sparse_ratio = clamp(sparse_row_ratio * clamp(active_column_count / 20.0))
    dense_penalty = clamp((average_non_empty_per_row - 8.0) / 8.0)
    occupancy_ratio = total_non_empty / (scanned_rows * scanned_columns) if scanned_columns else 0.0
    lowered_sheet_name = ws.title.strip().lower()
    positive_hits = [token for token in positive_name_tokens if token in lowered_sheet_name]
    negative_hits = [token for token in negative_name_tokens if token in lowered_sheet_name]
    sheet = {
        "sheet_name": ws.title,
        "scanned_rows": scanned_rows,
        "scanned_columns": scanned_columns,
        "active_column_count": active_column_count,
        "active_columns": active_columns,
        "average_non_empty_per_row": average_non_empty_per_row,
        "text_cell_ratio": text_cell_ratio,
        "numeric_cell_ratio": numeric_cell_ratio,
        "styled_cell_ratio": styled_cell_ratio,
        "header_like_row_ratio": header_like_row_ratio,
        "sparse_row_ratio": sparse_row_ratio,
        "instruction_signal": instruction_signal,
        "lookup_signal": lookup_signal,
        "template_shape": template_shape,
        "header_focus": header_focus,
        "wide_sparse_ratio": wide_sparse_ratio,
        "dense_penalty": dense_penalty,
        "occupancy_ratio": occupancy_ratio,
        "candidate_header_rows": candidate_header_rows[:3],
        "first_header_row": first_header_row if first_header_row <= scanned_rows else 0,
        "name_positive_ratio": len(positive_hits) / max(1, len(positive_name_tokens)),
        "name_negative_ratio": len(negative_hits) / max(1, len(negative_name_tokens)),
        "name_positive_hits": positive_hits[:3],
        "name_negative_hits": negative_hits[:3],
    }
    return {"sheet": sheet, "rows": rows}
def _score_additive(features: dict[str, float], weights: dict[str, float]) -> dict[str, Any]:
    components: dict[str, float] = {}
    for key, weight in weights.items():
        components[key] = round(float(features.get(key, 0.0)) * float(weight), 6)
    score = clamp(sum(components.values()))
    return {
        "score": round(score, 6),
        "components": components,
        "features": {key: round(float(features.get(key, 0.0)), 6) for key in weights},
    }
def score_sheet(feature_dump: dict[str, Any], weights: dict[str, float]) -> dict[str, Any]:
    sheet = feature_dump["sheet"]
    sheet_features = {
        "column_breadth": clamp(sheet["active_column_count"] / 20.0),
        "density": clamp(sheet["average_non_empty_per_row"] / 8.0),
        "text_balance": clamp(sheet["text_cell_ratio"] * (1.0 - 0.5 * sheet["numeric_cell_ratio"])),
        "style_support": clamp(sheet["styled_cell_ratio"] * 2.0),
        "name_positive": sheet["name_positive_ratio"],
        "template_shape": sheet["template_shape"],
        "header_focus": sheet["header_focus"],
        "wide_sparse_ratio": sheet["wide_sparse_ratio"],
        "instruction_signal": sheet["instruction_signal"],
        "lookup_signal": sheet["lookup_signal"],
        "dense_penalty": sheet["dense_penalty"],
        "name_negative": sheet["name_negative_ratio"],
    }
    return _score_additive(sheet_features, weights)
def detect_header_row(
    feature_dump: dict[str, Any],
    top_k: int,
    weights: dict[str, float],
) -> dict[str, Any]:
    rows = feature_dump["rows"]
    scanned_rows = max(1, len(rows))
    candidates: list[dict[str, Any]] = []
    for idx, row in enumerate(rows):
        next_rows = rows[idx + 1 : idx + 3]
        next_avg_non_empty = (
            sum(next_row["non_empty_count"] for next_row in next_rows) / max(1, len(next_rows))
        )
        non_empty_count = row["non_empty_count"]
        style_ratio = clamp(row["styled_count"] / max(1.0, float(non_empty_count)))
        features = {
            "non_empty_norm": clamp(non_empty_count / 10.0),
            "string_ratio": row["string_ratio"],
            "short_text_ratio": row["short_text_ratio"],
            "unique_value_ratio": row["unique_value_ratio"],
            "style_ratio": style_ratio,
            "followed_by_tabular": clamp(next_avg_non_empty / max(3.0, float(non_empty_count))),
            "early_row_bias": clamp(1.0 - ((row["row_index"] - 1) / max(1.0, scanned_rows - 1.0))),
            "has_id_like_token": float(row["has_id_like_token"]),
            "has_name_like_token": float(row["has_name_like_token"]),
            "has_date_like_token": float(row["has_date_like_token"]),
            "has_code_like_token": float(row["has_code_like_token"]),
            "numeric_ratio": row["numeric_ratio"],
            "long_text_ratio": row["long_text_ratio"],
            "sparse_penalty": clamp((2.0 - non_empty_count) / 2.0),
        }
        score_info = _score_additive(features, weights)
        candidates.append(
            {
                "row": row["row_index"],
                "score": score_info["score"],
                "components": score_info["components"],
                "features": score_info["features"],
            }
        )
    candidates.sort(key=lambda item: (-item["score"], item["row"]))
    top_candidates = candidates[: max(1, top_k)]
    best = top_candidates[0]
    return {
        "best_row": int(best["row"]),
        "score": best["score"],
        "components": best["components"],
        "features": best["features"],
        "top_candidates": top_candidates,
    }
def detect_data_row(
    feature_dump: dict[str, Any],
    header_row: int,
    top_k: int,
    weights: dict[str, float],
) -> dict[str, Any]:
    rows = feature_dump["rows"]
    candidates = [row for row in rows if row["row_index"] > header_row]
    if not candidates:
        fallback = {"row": int(header_row + 1), "score": 0.0, "components": {}, "features": {}}
        return {
            "best_row": int(header_row + 1),
            "score": 0.0,
            "components": {},
            "features": {},
            "top_candidates": [fallback],
        }
    header_features = next((row for row in rows if row["row_index"] == header_row), None)
    header_columns = set(header_features["non_empty_columns"] if header_features else [])
    if not header_columns:
        header_columns = set(feature_dump["sheet"].get("active_columns", []))
    header_col_count = max(1, len(header_columns))
    first_non_empty_after_header = next(
        (row["row_index"] for row in candidates if row["non_empty_count"] >= 2),
        candidates[0]["row_index"],
    )
    span = max(1, candidates[-1]["row_index"] - (header_row + 1) + 1)
    rows_by_index = {row["row_index"]: row for row in rows}
    scored_rows: list[dict[str, Any]] = []
    for row in candidates:
        non_empty_count = row["non_empty_count"]
        row_columns = set(row["non_empty_columns"])
        overlap_with_header = len(row_columns & header_columns) / header_col_count
        style_ratio = clamp(row["styled_count"] / max(1.0, float(non_empty_count)))
        distance = row["row_index"] - header_row
        prev_non_empty = rows_by_index.get(row["row_index"] - 1, {}).get("non_empty_count", 0)
        features = {
            "overlap_with_header": clamp(overlap_with_header),
            "numeric_ratio": row["numeric_ratio"],
            "value_mix": clamp(1.0 - row["short_text_ratio"] * row["string_ratio"]),
            "early_after_header": clamp(1.0 - ((distance - 1) / span)),
            "offset_preference": clamp(1.0 - abs(distance - 2.5) / 3.5),
            "unstyled_ratio": clamp(1.0 - style_ratio),
            "unique_value_ratio": row["unique_value_ratio"],
            "first_non_empty_after_header": float(row["row_index"] == first_non_empty_after_header),
            "sparse_input": float(non_empty_count <= 1),
            "transition_from_dense": float(prev_non_empty >= 3 and non_empty_count <= 1),
            "header_like_text": clamp(row["string_ratio"] * row["short_text_ratio"]),
            "long_note_penalty": clamp(row["long_text_ratio"] * row["string_ratio"]),
            "dense_row_penalty": clamp((non_empty_count - 8.0) / 8.0),
        }
        score_info = _score_additive(features, weights)
        scored_rows.append(
            {
                "row": row["row_index"],
                "score": score_info["score"],
                "components": score_info["components"],
                "features": score_info["features"],
            }
        )
    scored_rows.sort(key=lambda item: (-item["score"], item["row"]))
    top_candidates = scored_rows[: max(1, top_k)]
    best = top_candidates[0]
    return {
        "best_row": int(best["row"]),
        "score": best["score"],
        "components": best["components"],
        "features": best["features"],
        "top_candidates": top_candidates,
    }
def select_best(
    workbooks: list[Any],
    max_rows: int = MAX_SCAN_ROWS_DEFAULT,
    knobs: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Select best workbook/sheet and return prediction + full feature/score dump."""
    settings = _resolve_settings(max_rows=max_rows, knobs=knobs)
    weights = settings["weights"]
    workbook_candidates: list[dict[str, Any]] = []
    for entry in workbooks:
        workbook_name: str | None = None
        workbook_obj: Any = None
        if isinstance(entry, tuple) and len(entry) >= 2:
            workbook_name, workbook_obj = str(entry[0]), entry[1]
        elif isinstance(entry, dict):
            workbook_name = str(entry.get("filename", ""))
            workbook_obj = entry.get("workbook")
        if not workbook_name or workbook_obj is None:
            continue
        visible_sheets = get_visible_sheets(workbook_obj)
        if not visible_sheets:
            continue
        sheet_candidates: list[dict[str, Any]] = []
        for ws in visible_sheets:
            feature_dump = extract_features(
                ws=ws,
                max_rows=settings["max_scan_rows"],
                max_columns=settings["max_scan_columns"],
                positive_name_tokens=settings["sheet_name_positive_tokens"],
                negative_name_tokens=settings["sheet_name_negative_tokens"],
            )
            sheet_score = score_sheet(feature_dump, weights=weights["sheet"])
            header_result = detect_header_row(
                feature_dump=feature_dump,
                top_k=settings["top_k"],
                weights=weights["header"],
            )
            data_result = detect_data_row(
                feature_dump=feature_dump,
                header_row=header_result["best_row"],
                top_k=settings["top_k"],
                weights=weights["data"],
            )
            selection_result = _score_additive(
                {
                    "sheet": sheet_score["score"],
                    "header": header_result["score"],
                    "data": data_result["score"],
                },
                weights["selection"],
            )
            sheet_candidates.append(
                {
                    "sheet_name": ws.title,
                    "sheet_score": sheet_score,
                    "header": header_result,
                    "data": data_result,
                    "selection": selection_result,
                    "feature_dump": feature_dump,
                }
            )
        if not sheet_candidates:
            continue
        sheet_candidates.sort(key=lambda item: (-item["selection"]["score"], item["sheet_name"].lower()))
        best_sheet = sheet_candidates[0]
        workbook_score = _score_additive(
            {
                "best_sheet_selection": best_sheet["selection"]["score"],
                "visible_sheet_bonus": clamp(len(visible_sheets) / 4.0),
            },
            weights["workbook"],
        )
        workbook_candidates.append(
            {
                "workbook_name": workbook_name,
                "workbook_score": workbook_score,
                "sheet_candidates": sheet_candidates,
                "best_sheet": best_sheet,
            }
        )
    if not workbook_candidates:
        raise ValueError("No visible sheets found in any valid workbook.")
    workbook_candidates.sort(key=lambda item: (-item["workbook_score"]["score"], item["workbook_name"].lower()))
    selected = workbook_candidates[0]
    best_sheet = selected["best_sheet"]
    confidence_score = _score_additive(
        {
            "workbook": selected["workbook_score"]["score"],
            "sheet": best_sheet["sheet_score"]["score"],
            "header": best_sheet["header"]["score"],
            "data": best_sheet["data"]["score"],
        },
        weights["confidence"],
    )
    workbook_top_candidates = [
        {
            "workbook": candidate["workbook_name"],
            "score": candidate["workbook_score"]["score"],
            "components": candidate["workbook_score"]["components"],
            "best_sheet": candidate["best_sheet"]["sheet_name"],
        }
        for candidate in workbook_candidates[: settings["top_k"]]
    ]
    sheet_top_candidates = [
        {
            "sheet": candidate["sheet_name"],
            "score": candidate["sheet_score"]["score"],
            "components": candidate["sheet_score"]["components"],
            "features": candidate["sheet_score"]["features"],
        }
        for candidate in selected["sheet_candidates"][: settings["top_k"]]
    ]
    prediction = {
        "sheet": best_sheet["sheet_name"],
        "header_row": int(best_sheet["header"]["best_row"]),
        "data_row": int(best_sheet["data"]["best_row"]),
    }
    return {
        "prediction": prediction,
        "selected_workbook": selected["workbook_name"],
        "selected_sheet": prediction["sheet"],
        "header_row": prediction["header_row"],
        "data_row": prediction["data_row"],
        "confidence": {
            "workbook": round(selected["workbook_score"]["score"], 4),
            "sheet": round(best_sheet["sheet_score"]["score"], 4),
            "header_row": round(best_sheet["header"]["score"], 4),
            "data_row": round(best_sheet["data"]["score"], 4),
            "overall": round(confidence_score["score"], 4),
        },
        "scores": {
            "workbook": {
                "score": selected["workbook_score"]["score"],
                "components": selected["workbook_score"]["components"],
                "top_candidates": workbook_top_candidates,
            },
            "sheet": {
                "score": best_sheet["sheet_score"]["score"],
                "components": best_sheet["sheet_score"]["components"],
                "features": best_sheet["sheet_score"]["features"],
                "top_candidates": sheet_top_candidates,
            },
            "header": best_sheet["header"],
            "data": best_sheet["data"],
            "selection": {
                "score": best_sheet["selection"]["score"],
                "components": best_sheet["selection"]["components"],
            },
        },
        "features": {
            "sheet": best_sheet["feature_dump"]["sheet"],
            "rows": best_sheet["feature_dump"]["rows"],
        },
    }
def auto_detect_config(
    zip_path: str,
    max_rows: int = MAX_SCAN_ROWS_DEFAULT,
    knobs: dict[str, Any] | None = None,
) -> dict[str, Any]:
    workbooks: list[tuple[str, Any]] = []
    for filename, file_bytes in extract_excel_files(zip_path):
        try:
            workbooks.append((filename, load_workbook(file_bytes)))
        except Exception:
            continue
    if not workbooks:
        raise ValueError("Failed to load any valid .xlsx/.xlsm workbook from ZIP.")
    result = select_best(workbooks=workbooks, max_rows=max_rows, knobs=knobs)
    result["zip_path"] = zip_path
    return result
def auto_detect_config_from_excel(
    excel_path: str,
    max_rows: int = MAX_SCAN_ROWS_DEFAULT,
    knobs: dict[str, Any] | None = None,
) -> dict[str, Any]:
    if not os.path.exists(excel_path):
        raise ValueError(f"Invalid Excel file path: {excel_path}")
    if not excel_path.lower().endswith(SUPPORTED_EXCEL_EXTENSIONS):
        supported = ", ".join(SUPPORTED_EXCEL_EXTENSIONS)
        raise ValueError(f"Only {supported} is supported in this function: {excel_path}")
    with open(excel_path, "rb") as f:
        workbook = load_workbook(f.read())
    result = select_best(
        workbooks=[(os.path.basename(excel_path), workbook)],
        max_rows=max_rows,
        knobs=knobs,
    )
    result["zip_path"] = excel_path
    return result
